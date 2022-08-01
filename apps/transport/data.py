import datetime
import os
import re

import openpyxl
import pandas as pd
from django.db import IntegrityError

from apps.models import Person
from apps.transport.models import BusRider, BusRoute
from qux.utils.phone import phone_number


def import_excel():
    filepath = "/Users/vishal/Downloads/transportroutelist/"
    files = os.listdir(filepath)
    failed = []
    dflist = []
    for filename in files:
        f = os.path.join(filepath, filename)
        df = pd.read_excel(
            f,
            header=None,
            names=[
                "x",
                "name",
                "grade",
                "phone",
                "address",
                "area",
                "time",
                "filename",
            ],
        )
        df["filename"] = filename
        df.name = df.name.str.strip()
        row = df.index[(df.name.str.endswith("Name") == True)]
        if row.empty:
            failed.append(df)
        else:
            df = df.loc[row[0] :]
            df = df[df.columns[1:8]]
            dflist.append(df)


def import_better():
    filepath = "/Users/vishal/Downloads/revisedtransportrouteslist/"
    files = os.listdir(filepath)
    routes = {}
    records = []
    for filenum, filename in enumerate(files):
        f = os.path.join(filepath, filename)
        wb = openpyxl.load_workbook(f)
        ws = wb.active

        match = re.search(r"Route (\d+\w) .*", filename)
        if match:
            route_number = match[1]
        else:
            continue

        route_name = ws.cell(1, 1).value
        route_details = ws.cell(2, 1).value
        bus_parent = ws.cell(3, 1).value
        attendant = ws.cell(4, 1).value
        route = {
            "name": route_name,
            "route": route_details,
            "bus_parent": bus_parent,
            "attendant": attendant,
            "object": BusRoute.objects.get(route=route_number),
        }
        routes[filenum] = route

        for r in range(6, ws.max_row + 1):
            phone = str(ws.cell(r, 4).value)
            phone = phone.split(",")[0].split("/")[0]
            phone = phone_number(phone)
            record = {
                "route": filenum,
                "name": ws.cell(r, 2).value,
                "grade": ws.cell(r, 3).value,
                "phone": phone,
                "address": ws.cell(r, 5).value,
                "area": ws.cell(r, 6).value,
                "time": ws.cell(r, 7).value,
            }
            records.append(record)

    # for route in routes:
    #     routes[route]["object"] = BusRoute.create_record(routes[route])

    robj = BusRoute.objects.get(id=28)

    for r in records:
        for x in ["rider", "address", "area"]:
            if hasattr(r.get(x, None), "strip"):
                pieces = [x.strip().upper() for x in str(r[x]).strip().split(",")]
                r[x] = ", ".join(pieces)

        person = Person.create_record(r)
        if person is None:
            continue

        route = routes[r["route"]]["object"]

        pickup_time = str(r["time"]).replace(".", ":").split(":")
        if all([x.isdigit() for x in pickup_time]):
            pickup_time = [int(x) for x in pickup_time]
        else:
            continue

        rparams = {
            "rider_id": person.id,
            "route_id": route.id,
            "area": r["area"],
            "time": datetime.time(*pickup_time),
        }

        try:
            rider, created = BusRider.objects.get_or_create(**rparams)
        except IntegrityError:
            continue

    return routes, records
